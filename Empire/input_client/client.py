from pathlib import Path

import openpyxl
import pandas as pd

from Empire.input_client.sheets_structure import sheets


class BaseClient:
    DEFAULT_SKIPROWS = None
    DEFAULT_USECOLS = None
    DEFAULT_STARTROW = None

    def _read_from_sheet(
        self, file_path: Path, sheet_name: str, **kwargs
    ) -> pd.DataFrame:
        """
        Read data from a specific sheet.

        :param file_path: Path to the Excel file.
        :param sheet_name: Name of the sheet to read from.
        :returns: Data from the sheet.
        """
        skiprows = kwargs.pop("skiprows", self.DEFAULT_SKIPROWS)
        usecols = kwargs.pop("usecols", self.DEFAULT_USECOLS)
        return pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            engine=self.engine,
            skiprows=skiprows,
            usecols=usecols,
            **kwargs,
        )

    def _write_to_sheet(
        self, df: pd.DataFrame, file_path: Path, sheet_name: str, **kwargs
    ):
        """
        Write data to a specific sheet.

        :param df: Data to the sheet
        :param file_path: Path to the Excel file.
        :param sheet_name: Name of the sheet to write to.
        """
        startrow = kwargs.pop("startrow", self.DEFAULT_STARTROW)
        with pd.ExcelWriter(
            file_path, engine=self.engine, mode="a", if_sheet_exists="replace"
        ) as writer:
            df.to_excel(
                writer, sheet_name=sheet_name, index=False, startrow=startrow, **kwargs
            )

    def validate(self):
        """Validate if the Excel file has the expected sheet names."""
        name = self.__class__.__name__.split("Client", maxsplit=1)[0]
        wb = openpyxl.load_workbook(self.file)
        if set(wb.sheetnames) != set(sheets[name]):
            raise ValueError(
                f"Sheetnames in {self.file} dont match expected sheet names for {name}."
                f"Expected: {sheets[name]}, Found: {wb.sheetnames}"
            )


class SetsClient(BaseClient):
    DEFAULT_SKIPROWS = 0

    def __init__(self, file, engine: str = "openpyxl"):
        self.file = file
        self.engine = engine

        self.validate()

    def get_nodes(self):
        return self._read_from_sheet(self.file, "Nodes")

    def set_nodes(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "Nodes")

    def get_offshore_nodes(self):
        return self._read_from_sheet(self.file, "OffshoreNodes")

    def set_offshore_nodes(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "OffshoreNodes")

    def get_technology(self):
        return self._read_from_sheet(self.file, "Technology")

    def set_technology(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "Technology")

    def get_directional_lines(self):
        return self._read_from_sheet(self.file, "DirectionalLines", skiprows=2)

    def set_directional_lines(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "DirectionalLines", startrow=2)


class GeneratorClient(BaseClient):
    DEFAULT_SKIPROWS = 2
    DEFAULT_STARTROW = 2

    def __init__(self, file: Path, engine: str = "openpyxl"):
        self.file = file
        self.engine = engine

        self.validate()

    def get_captial_costs(self):
        return self._read_from_sheet(self.file, "CapitalCosts")

    def set_captial_costs(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "CapitalCosts")

    def get_fixed_om_costs(self):
        return self._read_from_sheet(self.file, "FixedOMCosts")

    def set_fixed_om_costs(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "FixedOMCosts")

    def get_variable_om_costs(self):
        return self._read_from_sheet(self.file, "VariableOMCosts")

    def set_variable_om_costs(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "VariableOMCosts")

    def get_fuel_costs(self):
        return self._read_from_sheet(self.file, "FuelCosts")

    def set_fuel_costs(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "FuelCosts")

    def get_efficiency(self):
        return self._read_from_sheet(self.file, "Efficiency")

    def set_efficiency(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "Efficiency")

    def get_ref_initial_capacity(self):
        return self._read_from_sheet(self.file, "RefInitialCap")

    def set_ref_initial_capacity(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "RefInitialCap")

    def get_scale_factor_initial_capacity(self):
        return self._read_from_sheet(self.file, "ScaleFactorInitialCap")

    def set_scale_factor_initial_capacity(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "ScaleFactorInitialCap")

    def get_max_built_capacity(self):
        return self._read_from_sheet(self.file, "MaxBuiltCapacity")

    def set_max_built_capacity(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "MaxBuiltCapacity")

    def get_max_installed_capacity(self):
        return self._read_from_sheet(
            self.file, "MaxInstalledCapacity", usecols=[0, 1, 2]
        )

    def set_max_installed_capacity(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "MaxInstalledCapacity")


class NodeClient(BaseClient):
    DEFAULT_STARTROW = 2
    DEFAULT_SKIPROWS = 2

    def __init__(self, file: Path, engine: str = "openpyxl"):
        self.file = file
        self.engine = engine

        self.validate()

    def get_electric_annual_demand(self):
        return self._read_from_sheet(
            self.file, "ElectricAnnualDemand", usecols=[0, 1, 2]
        )

    def set_electric_annual_demand(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "ElectricAnnualDemand")

    def get_node_lost_load_cost(self):
        return self._read_from_sheet(self.file, "NodeLostNodeCost", usecols=[0, 1, 2])

    def set_node_lost_load_cost(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "NodeLostNodeCost")

    def get_hydro_generators_max_annual_production(self):
        return self._read_from_sheet(
            self.file, "HydroGenMaxAnnualProduction", usecols=[0, 1]
        )

    def set_hydro_generators_max_annual_production(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "HydroGenMaxAnnualProduction")


class TransmissionClient(BaseClient):
    DEFAULT_SKIPROWS = 2
    DEFAULT_STARTROW = 2

    def __init__(self, file: Path, engine: str = "openpyxl"):
        self.file = file
        self.engine = engine

        self.validate()

    def get_line_efficiency(self):
        return self._read_from_sheet(self.file, "lineEfficiency", usecols=[0, 1, 2])

    def set_line_efficiency(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "lineEfficiency")

    def get_max_built_capacity(self):
        return self._read_from_sheet(
            self.file, "MaxBuiltCapacity", usecols=[0, 1, 2, 3]
        )

    def set_max_built_capacity(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "MaxBuiltCapacity")

    def get_length(self):
        return self._read_from_sheet(self.file, "Length", usecols=[0, 1, 2])

    def set_length(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "Length")

    def get_type_capital_cost(self):
        return self._read_from_sheet(self.file, "TypeCapitalCost", usecols=[0, 1, 2])

    def set_type_capital_cost(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "TypeCapitalCost")

    def get_type_fixed_om_cost(self):
        return self._read_from_sheet(self.file, "TypeFixedOMCost", usecols=[0, 1, 2])

    def set_type_fixed_om_cost(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "TypeFixedOMCost")

    def get_initial_capacity(self):
        return self._read_from_sheet(self.file, "InitialCapacity", usecols=[0, 1, 2, 3])

    def set_initial_capacity(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "InitialCapacity")

    def get_max_install_capacity_raw(self):
        return self._read_from_sheet(
            self.file, "MaxInstallCapacityRaw", usecols=[0, 1, 2, 3]
        )

    def set_max_install_capacity_raw(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "MaxInstallCapacityRaw")

    def get_lifetime(self):
        return self._read_from_sheet(self.file, "Lifetime", usecols=[0, 1, 2])

    def set_lifetime(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "Lifetime")


class StorageClient(BaseClient):
    DEFAULT_STARTROW = 2
    DEFAULT_SKIPROWS = 2

    def __init__(self, file, engine: str = "openpyxl"):
        self.file = file
        self.engine = engine

        self.validate()

    def get_initial_power_capacity(self):
        return self._read_from_sheet(
            self.file, "InitialPowerCapacity", usecols=[0, 1, 2, 3]
        )

    def set_initial_power_capacity(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "InitialPowerCapacity")

    def get_power_capital_cost(self):
        return self._read_from_sheet(self.file, "PowerCapitalCost", usecols=[0, 1, 2])

    def set_power_capital_cost(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "PowerCapitalCost")

    def get_power_max_built_capacity(self):
        return self._read_from_sheet(
            self.file, "PowerMaxBuiltCapacity", usecols=[0, 1, 2, 3]
        )

    def set_power_max_built_capacity(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "PowerMaxBuiltCapacity")

    def getEnergyCapitalCost(self):
        return self._read_from_sheet(self.file, "EnergyCapitalCost", usecols=[0, 1, 2])

    def setEnergyCapitalCost(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "EnergyCapitalCost")

    def get_energy_initial_capacity(self):
        return self._read_from_sheet(
            self.file, "EnergyInitialCapacity", usecols=[0, 1, 2, 3]
        )

    def set_energy_initial_capacity(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "EnergyInitialCapacity")

    def get_energy_max_built_capacity(self):
        return self._read_from_sheet(
            self.file, "EnergyMaxBuiltCapacity", usecols=[0, 1, 2, 3]
        )

    def set_energy_max_built_capacity(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "EnergyMaxBuiltCapacity")

    def get_energy_max_installed_capacity(self):
        return self._read_from_sheet(
            self.file, "EnergyMaxInstalledCapacity", usecols=[0, 1, 2]
        )

    def set_energy_max_installed_capacity(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "EnergyMaxInstalledCapacity")

    def get_power_max_installed_capacity(self):
        return self._read_from_sheet(
            self.file, "PowerMaxInstalledCapacity", usecols=[0, 1, 2]
        )

    def set_power_max_installed_capacity(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "PowerMaxInstalledCapacity")

    def get_storage_initial_energy_level(self):
        return self._read_from_sheet(
            self.file, "StorageInitialEnergyLevel", usecols=[0, 1]
        )

    def set_storage_initial_energy_level(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "StorageInitialEnergyLevel")

    def get_storage_charge_efficiency(self):
        return self._read_from_sheet(self.file, "StorageChargeEff", usecols=[0, 1])

    def set_storage_charge_efficiency(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "StorageChargeEff")

    def get_storage_discharge_efficiency(self):
        return self._read_from_sheet(self.file, "StorageDischargeEff", usecols=[0, 1])

    def set_storage_discharge_efficiency(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "StorageDischargeEff")

    def get_storage_power_to_energt(self):
        return self._read_from_sheet(self.file, "StoragePowToEnergy", usecols=[0, 1])

    def set_storage_power_to_energt(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "StoragePowToEnergy")

    def get_lifetime(self):
        return self._read_from_sheet(self.file, "Lifetime", usecols=[0, 1])

    def set_lifetime(self, df: pd.DataFrame):
        self._write_to_sheet(df, self.file, "Lifetime")


class EmpireInputClient:
    """
    A unified interface (facade) for accessing various input clients related to the
    Empire dataset.

    Attributes:
    - dataset_path (Path): The base directory containing all datasets.
    - engine (str): The engine for reading/writing Excel files. Default is "openpyxl".
    - sets (SetsClient): Client for managing 'Sets' data.
    - generator (GeneratorClient): Client for managing 'Generator' data.
    - nodes (NodeClient): Client for managing 'Node' data.
    - transmission (TransmissionClient): Client for managing 'Transmission' data.
    - storage (StorageClient): Client for managing 'Storage' data.
    """

    def __init__(self, dataset_path: Path, engine: str = "openpyxl"):
        """
        Initialize an EmpireInputClient with the given dataset path and engine.

        :param dataset_path: Base directory containing all datasets.
        :param engine: Engine to use for reading/writing Excel files. Default is "openpyxl".
        """
        self.dataset_path = dataset_path
        self.engine = engine

        self.sets = SetsClient(dataset_path / "Sets.xlsx")
        self.generator = GeneratorClient(dataset_path / "Generator.xlsx")
        self.nodes = NodeClient(dataset_path / "Node.xlsx")
        self.transmission = TransmissionClient(dataset_path / "Transmission.xlsx")
        self.storage = StorageClient(dataset_path / "Storage.xlsx")


# Example usage:
if __name__ == "__main__":
    # %%

    dataset_path = Path.cwd() / "Data handler/europe_v51"
    client = EmpireInputClient(dataset_path)
    # client.get_nodes()

    # df = client.get_directional_lines()
    # client.set_directional_lines(df)

    df = client.generator.get_captial_costs()
    df = client.generator.get_max_installed_capacity()

    df = client.transmission.get_lifetime()
